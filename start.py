#!/bin/env python3
"""Import Library"""
import selenium
from selenium import webdriver
import random
import chromedriver_autoinstaller
from time import sleep
from fake_useragent import UserAgent
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from openpyxl import load_workbook
from colorama import Style, Fore
from selenium.webdriver.common.proxy import Proxy, ProxyType

LINKS = [] #Array with links

"""LOGIN AND PASSWORD TO EXCEL"""
def load_data(login,password):
    wb = load_workbook('datafile.xlsx')
    sheet = wb['Sheet1']
    sheet.cell(row = 1, column = 1).value = 'Login'
    sheet.cell(row = 1,column = 2).value = 'Password'
    rw = sheet.max_row + 1
    sheet.cell(row = rw, column = 1).value = str(login)
    sheet.cell(row = rw, column = 2).value = str(password)
    wb.save('datafile.xlsx')
    print(Fore.GREEN + "Запись логина и пароля успешно произвелась..", Style.RESET_ALL)

"""APPEND ARRAY FROM links.txt"""
def load_links():
    with open('links.txt', 'r') as f:
        for el in f:
            LINKS.append(el.replace('/n', ''))
    print(Fore.GREEN + "Добавление ссылок успешно..", Style.RESET_ALL)
            



"""MAIN CLASS AUTOMIZATION"""
class Main_class:
    
    """Initialization"""
    def __init__(self, link):
        self.link = link # INITIALIZATE LINK
        self.email_url = 'https://generator.email/' #LINK TO GENERATE EMAIL 
        self.pornhub_url = 'https://rt.pornhub.com/signup' #PORNHUB EMAIL
    
    """Create brosweser"""
    def make_browser(self):
        opts = Options() #Create Options
        prox = Proxy()
        prox.proxy_type = ProxyType.MANUAL
        prox.http_proxy = "78.47.16.54:80"

        capabilities = webdriver.DesiredCapabilities.CHROME
        prox.add_to_capabilities(capabilities)
        ua = UserAgent() #Fake User Agent
        user_agent = ua.random #random user agent
        opts.add_argument(f'user-agent={user_agent}') #add argument
        self.driver = webdriver.Chrome('windows/chromedriver.exe',
                chrome_options=opts,
                desired_capabilities=capabilities) #Create Browser with Options
        print(Fore.GREEN + "Создание браузера завершилось!", Style.RESET_ALL)


    """Login and Password Generator"""
    def generate_login(self):
        chars = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890'
        self.password = ''
        for i in range(1,16):
            self.password += random.choice(chars)
        bg = self.password + '1'
        load_data(self.password, bg)
        print(Fore.GREEN + "Логин и пароль успешно создались!", Style.RESET_ALL)
    
    """Create Mail"""
    def make_mail(self):
        self.driver.get(self.email_url) #TO GENERATE EMAIL 
        username = self.driver.find_element_by_id('userName').get_attribute('value') #mail
        domain = self.driver.find_element_by_id('domainName2').get_attribute('value') #domain
        self.alls = username + "@" + domain #full adress: template@domain.com
        print(Fore.GREEN + "Создание почты завершилось!", Style.RESET_ALL)

    """Create Account in pornhub"""
    def create_account(self):
        self.driver.get(self.pornhub_url) # to pornhub
        email = self.driver.find_element_by_name('email').send_keys(self.alls) #send mail
        username = self.driver.find_element_by_css_selector('#registerWrapper > div > div.formWrapper > v-create-account-form > form > section > input[type=text]:nth-child(5)').send_keys(self.password) #send login
        password = self.driver.find_element_by_css_selector('#registerWrapper > div > div.formWrapper > v-create-account-form > form > section > input[type=password]:nth-child(7)').send_keys(self.password + str(1)) #send password
        sleep(2) #TIME TO SLEEP
        try:
            button = self.driver.find_element_by_css_selector('#acceptCookie').click() #skip cookie
        except:
            pass
        button = self.driver.find_element_by_css_selector('#registerWrapper > div > div.formWrapper > v-create-account-form > form > section > input.buttonBase.orangeButton.big').click() #register button 
        sleep(2)
        
    def check_mail(self):
        sleep(15) #wait message
        print(self.email_url + '/' + self.alls)
        def validate():
            global the_confirmation
            self.driver.get(self.email_url + '/' + self.alls)
            button1 = self.driver.find_element_by_css_selector('#refresh > button').click()
            the_confirmation = self.driver.find_element_by_css_selector('#email-table > div.e7m.row.list-group-item > div.e7m.col-md-12.ma1 > div.e7m.mess_bodiyy > table.table > tbody > tr:nth-child(8) > td > p > a')
        
        for i in range(1,5):
            try:
                validate()
            except:
                continue
        try:
            self.driver.get(the_confirmation.text)

            print(Fore.GREEN + "Регистрация успешно завершилась!", Style.RESET_ALL)
        except:
            self.view()
        
    """VIEW VIDEOS"""
    def view(self):
        for link in LINKS:
            self.driver.get(link)
            try:
                skip = self.driver.find_element_by_css_selector('#acceptCookie').click()
            except:
                pass
            sleep(4)
            button = self.driver.find_element_by_xpath("/html/body/div[7]/div/div[5]/div[1]/div[1]/div[1]/div[4]/div[3]/div[1]").click() #like
            sleep(3)
        try:
            self.driver.close()
        except:
            pass

    def closer(self):
        for i in range(1,555):
            try:
                self.driver.quit()
            except:
                break
def main():
    load_links()
    root = Main_class('fdf')
    root.make_browser()
    root.generate_login()
    root.make_mail()
    root.create_account()
    root.check_mail()


if __name__ == "__main__":
        main()
        
        #except:
            #root1 = Main_class('as')
            #root1.closer()
